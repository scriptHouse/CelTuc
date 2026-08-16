from decimal import Decimal

from django.test import TestCase
from rest_framework.test import APIClient

from productos.models import CategoriaProducto, Producto
from usuarios.models import Permiso, Rol, Usuario

from .models import (
    MovimientoStock,
    StockProducto,
    Sucursal,
    Venta,
    aplicar_ajuste,
    aplicar_transferencia,
    registrar_venta,
)


def _producto(nombre='Fuente 20W test', **kwargs):
    categoria, _ = CategoriaProducto.objects.get_or_create(nombre='Categoria test')
    return Producto.objects.create(categoria=categoria, nombre=nombre, **kwargs)


class OperacionesStockTests(TestCase):
    """La logica pura: ajustes, pisos en 0 y transferencias."""

    def setUp(self):
        self.producto = _producto()
        self.solar = Sucursal.objects.create(nombre='Solar test', orden=1)
        self.centro = Sucursal.objects.create(nombre='Centro test', orden=2)

    def test_ajuste_crea_fila_y_movimiento(self):
        fila, mov = aplicar_ajuste(self.producto, self.solar, delta=5)
        self.assertEqual(fila.cantidad, 5)
        self.assertEqual(mov.tipo, MovimientoStock.Tipo.INGRESO)
        self.assertEqual(mov.delta, 5)
        self.assertEqual(mov.resultante, 5)

    def test_cantidad_fija_calcula_delta(self):
        aplicar_ajuste(self.producto, self.solar, delta=10)
        fila, mov = aplicar_ajuste(self.producto, self.solar, cantidad=4)
        self.assertEqual(fila.cantidad, 4)
        self.assertEqual(mov.delta, -6)
        self.assertEqual(mov.tipo, MovimientoStock.Tipo.EGRESO)

    def test_no_baja_de_cero(self):
        aplicar_ajuste(self.producto, self.solar, delta=2)
        from django.core.exceptions import ValidationError
        with self.assertRaises(ValidationError):
            aplicar_ajuste(self.producto, self.solar, delta=-3)
        fila = StockProducto.objects.get(producto=self.producto, sucursal=self.solar)
        self.assertEqual(fila.cantidad, 2)  # no cambio nada

    def test_delta_cero_no_registra_movimiento(self):
        aplicar_ajuste(self.producto, self.solar, delta=3)
        antes = MovimientoStock.objects.count()
        aplicar_ajuste(self.producto, self.solar, cantidad=3)
        self.assertEqual(MovimientoStock.objects.count(), antes)

    def test_transferencia_mueve_y_deja_dos_movimientos(self):
        aplicar_ajuste(self.producto, self.solar, delta=10)
        salida, entrada = aplicar_transferencia(self.producto, self.solar, self.centro, 4)
        self.assertEqual(salida.cantidad, 6)
        self.assertEqual(entrada.cantidad, 4)
        movs = MovimientoStock.objects.filter(tipo=MovimientoStock.Tipo.TRANSFERENCIA)
        self.assertEqual(movs.count(), 2)
        self.assertEqual(sorted(m.delta for m in movs), [-4, 4])

    def test_ajuste_limpia_sin_dato(self):
        # Una fila "(no informado)" deja de estarlo en cuanto alguien carga
        # una cantidad real — incluso si lo que carga es 0 (conto y habia 0).
        StockProducto.objects.create(producto=self.producto, sucursal=self.solar, sin_dato=True)
        fila, mov = aplicar_ajuste(self.producto, self.solar, cantidad=0)
        self.assertFalse(fila.sin_dato)
        self.assertIsNone(mov)  # el 0 explicito no genera movimiento
        StockProducto.objects.create(producto=self.producto, sucursal=self.centro, sin_dato=True)
        fila, mov = aplicar_ajuste(self.producto, self.centro, delta=3)
        self.assertFalse(fila.sin_dato)
        self.assertEqual(mov.tipo, MovimientoStock.Tipo.INGRESO)

    def test_transferencia_sin_stock_falla_atomica(self):
        aplicar_ajuste(self.producto, self.solar, delta=2)
        from django.core.exceptions import ValidationError
        with self.assertRaises(ValidationError):
            aplicar_transferencia(self.producto, self.solar, self.centro, 5)
        self.assertEqual(
            StockProducto.objects.get(producto=self.producto, sucursal=self.solar).cantidad, 2,
        )
        self.assertFalse(
            StockProducto.objects.filter(producto=self.producto, sucursal=self.centro).exists(),
        )


class VentasTests(TestCase):
    """La venta de mostrador: descuenta stock, registra kardex, todo o nada."""

    def setUp(self):
        self.fuente = _producto('Fuente venta test')
        self.cable = _producto('Cable venta test')
        self.solar = Sucursal.objects.create(nombre='Solar test', orden=1)
        aplicar_ajuste(self.fuente, self.solar, delta=10)
        aplicar_ajuste(self.cable, self.solar, delta=2)

        from usuarios.models import Permiso, Rol, Usuario
        rol = Rol.objects.create(nombre='Mostrador ventas test')
        rol.permisos.set(Permiso.objects.filter(codigo='ver_inventario'))
        self.empleado = Usuario.objects.create_user(
            email='vtas@celtuc.test', username='vtas.inv', password='x', rol=rol,
        )
        self.cliente = APIClient()
        self.cliente.force_authenticate(self.empleado)

    def test_registrar_venta_descuenta_y_totaliza(self):
        venta = registrar_venta(
            self.solar,
            [(self.fuente, 2, Decimal('38800')), (self.cable, 1, Decimal('11600'))],
            forma_pago='efectivo',
            usuario=self.empleado,
        )
        self.assertEqual(venta.total, Decimal('89200'))
        self.assertEqual(StockProducto.objects.get(producto=self.fuente, sucursal=self.solar).cantidad, 8)
        self.assertEqual(StockProducto.objects.get(producto=self.cable, sucursal=self.solar).cantidad, 1)
        movs = MovimientoStock.objects.filter(tipo=MovimientoStock.Tipo.VENTA)
        self.assertEqual(movs.count(), 2)
        self.assertTrue(all(m.nota == f'Venta #{venta.pk}' for m in movs))

    def test_venta_sin_stock_no_registra_nada(self):
        from django.core.exceptions import ValidationError
        with self.assertRaises(ValidationError):
            registrar_venta(
                self.solar,
                [(self.fuente, 1, Decimal('38800')), (self.cable, 5, Decimal('11600'))],
            )
        # Atomica: ni la venta, ni items, ni el descuento de la fuente.
        self.assertEqual(Venta.objects.count(), 0)
        self.assertEqual(StockProducto.objects.get(producto=self.fuente, sucursal=self.solar).cantidad, 10)
        self.assertEqual(MovimientoStock.objects.filter(tipo=MovimientoStock.Tipo.VENTA).count(), 0)

    def test_venta_con_un_solo_medio_deja_su_fila_de_pago(self):
        """Toda venta queda con al menos un pago: el dato es uniforme."""
        venta = registrar_venta(
            self.solar, [(self.fuente, 1, Decimal('38800'))], forma_pago='tarjeta',
        )
        pagos = list(venta.pagos.all())
        self.assertEqual(len(pagos), 1)
        self.assertEqual((pagos[0].medio, pagos[0].monto), ('tarjeta', Decimal('38800')))

    def test_venta_con_pagos_divididos(self):
        venta = registrar_venta(
            self.solar,
            [(self.fuente, 1, Decimal('38800')), (self.cable, 1, Decimal('11600'))],  # 50400
            pagos=[
                {'medio': 'efectivo', 'monto': Decimal('20400')},
                {'medio': 'transferencia', 'monto': Decimal('30000')},
            ],
        )
        self.assertEqual(venta.total, Decimal('50400'))
        self.assertEqual(
            {p.medio: p.monto for p in venta.pagos.all()},
            {'efectivo': Decimal('20400'), 'transferencia': Decimal('30000')},
        )
        # El medio principal (el de mayor monto) es el que ven los reportes.
        self.assertEqual(venta.forma_pago, 'transferencia')

    def test_dos_partes_de_factura_c_con_cuentas_distintas_no_se_fusionan(self):
        """Cada parte facturada es una factura aparte: con su propia cuenta."""
        from facturacion.models import Emisor

        uno = Emisor.objects.create(
            nombre='Mono Uno test', condicion='monotributista', cuit='20111111112',
        )
        dos = Emisor.objects.create(
            nombre='Mono Dos test', condicion='monotributista', cuit='20111111113',
        )
        venta = registrar_venta(
            self.solar,
            [(self.fuente, 1, Decimal('38800'))],
            pagos=[
                {'medio': 'efectivo', 'facturacion': 'factura_c',
                 'emisor': uno, 'monto': Decimal('18800')},
                {'medio': 'efectivo', 'facturacion': 'factura_c',
                 'emisor': dos, 'monto': Decimal('20000')},
            ],
        )
        pagos = list(venta.pagos.all())
        self.assertEqual(len(pagos), 2)  # mismo medio y facturacion, cuentas distintas
        self.assertEqual(
            {p.emisor_id: p.monto for p in pagos},
            {uno.pk: Decimal('18800'), dos.pk: Decimal('20000')},
        )

    def test_pagos_repetidos_se_suman_en_una_sola_parte(self):
        venta = registrar_venta(
            self.solar,
            [(self.fuente, 1, Decimal('38800'))],
            pagos=[
                {'medio': 'efectivo', 'monto': Decimal('8800')},
                {'medio': 'efectivo', 'monto': Decimal('30000')},
            ],
        )
        pagos = list(venta.pagos.all())
        self.assertEqual(len(pagos), 1)
        self.assertEqual(pagos[0].monto, Decimal('38800'))

    def test_diferencia_de_un_centavo_se_ajusta_sola(self):
        """Un redondeo de precios no puede hacer rebotar la venta."""
        venta = registrar_venta(
            self.solar,
            [(self.fuente, 1, Decimal('38800.005'))],
            pagos=[
                {'medio': 'efectivo', 'monto': Decimal('18800')},
                {'medio': 'tarjeta', 'monto': Decimal('20000')},
            ],
        )
        self.assertEqual(
            sum((p.monto for p in venta.pagos.all()), Decimal('0')), venta.total,
        )

    def test_pagos_que_no_suman_el_total_no_registran_nada(self):
        from django.core.exceptions import ValidationError
        with self.assertRaises(ValidationError):
            registrar_venta(
                self.solar,
                [(self.fuente, 1, Decimal('38800'))],
                pagos=[{'medio': 'efectivo', 'monto': Decimal('1000')}],
            )
        # Atomica: ni venta, ni pagos, ni descuento de stock.
        self.assertEqual(Venta.objects.count(), 0)
        self.assertEqual(
            StockProducto.objects.get(producto=self.fuente, sucursal=self.solar).cantidad, 10,
        )

    def test_api_venta_con_pagos_divididos(self):
        r = self.cliente.post('/api/inventario/ventas/', {
            'sucursal': self.solar.id,
            'items': [{'producto': self.fuente.id, 'cantidad': 1, 'precio_unitario': 38800}],
            'pagos': [
                {'medio': 'efectivo', 'monto': 18800},
                {'medio': 'tarjeta', 'monto': 20000},
            ],
        }, format='json')
        self.assertEqual(r.status_code, 201)
        self.assertEqual(len(r.data['pagos']), 2)
        self.assertEqual(sum(float(p['monto']) for p in r.data['pagos']), 38800)

    def test_api_pagos_que_no_cierran_dan_400_legible(self):
        r = self.cliente.post('/api/inventario/ventas/', {
            'sucursal': self.solar.id,
            'items': [{'producto': self.fuente.id, 'cantidad': 1, 'precio_unitario': 38800}],
            'pagos': [{'medio': 'efectivo', 'monto': 100}],
        }, format='json')
        self.assertEqual(r.status_code, 400)
        self.assertIn('coincidir', r.data['detail'])

    def test_venta_confirmada_con_faltante_registra_y_deja_negativo(self):
        """La venta NUNCA se pierde: confirmado el faltante, el stock queda en rojo."""
        venta = registrar_venta(
            self.solar,
            [(self.cable, 5, Decimal('11600'))],  # hay 2, se venden 5
            permitir_faltante=True,
        )
        self.assertEqual(venta.total, Decimal('58000'))
        self.assertEqual(
            StockProducto.objects.get(producto=self.cable, sucursal=self.solar).cantidad, -3,
        )
        mov = MovimientoStock.objects.get(tipo=MovimientoStock.Tipo.VENTA)
        self.assertEqual(mov.delta, -5)
        self.assertEqual(mov.resultante, -3)

    def test_api_venta_con_faltante_confirmado_201(self):
        r = self.cliente.post('/api/inventario/ventas/', {
            'sucursal': self.solar.id,
            'permitir_faltante': True,
            'items': [{'producto': self.cable.id, 'cantidad': 5, 'precio_unitario': 11600}],
        }, format='json')
        self.assertEqual(r.status_code, 201)
        self.assertEqual(
            StockProducto.objects.get(producto=self.cable, sucursal=self.solar).cantidad, -3,
        )

    def test_api_post_y_get(self):
        r = self.cliente.post('/api/inventario/ventas/', {
            'sucursal': self.solar.id,
            'forma_pago': 'tarjeta',
            'nota': 'cliente del 13',
            'items': [
                {'producto': self.fuente.id, 'cantidad': 1, 'precio_unitario': 38800},
            ],
        }, format='json')
        self.assertEqual(r.status_code, 201)
        self.assertEqual(float(r.data['total']), 38800)
        self.assertEqual(r.data['usuario'], 'vtas.inv')
        self.assertEqual(r.data['items'][0]['nombre'], 'Fuente venta test')
        r = self.cliente.get(f'/api/inventario/ventas/?sucursal={self.solar.id}')
        self.assertEqual(r.status_code, 200)
        self.assertEqual(len(r.data), 1)

    def test_api_sin_stock_400_legible(self):
        r = self.cliente.post('/api/inventario/ventas/', {
            'sucursal': self.solar.id,
            'items': [{'producto': self.cable.id, 'cantidad': 99, 'precio_unitario': 100}],
        }, format='json')
        self.assertEqual(r.status_code, 400)
        self.assertIn('stock suficiente', r.data['detail'])

    def test_api_requiere_permiso(self):
        from usuarios.models import Usuario
        pelado = Usuario.objects.create_user(
            email='pelado.v@celtuc.test', username='pelado.v', password='x',
        )
        cliente = APIClient()
        cliente.force_authenticate(pelado)
        self.assertEqual(cliente.get('/api/inventario/ventas/').status_code, 403)


class VentaServiciosTests(TestCase):
    """El mostrador no vende solo mercaderia: services e items libres.

    Lo importante: esos renglones se cobran igual pero NO tocan el stock, y la
    venta de solo productos sigue funcionando exactamente como siempre.
    """

    def setUp(self):
        from precios_service.models import ItemService, SeccionService

        self.fuente = _producto('Fuente service test')
        self.solar = Sucursal.objects.create(nombre='Solar service test', orden=1)
        aplicar_ajuste(self.fuente, self.solar, delta=10)
        seccion = SeccionService.objects.create(nombre='Baterias test')
        self.item_service = ItemService.objects.create(seccion=seccion, etiqueta='iPhone 13')

        from usuarios.models import Permiso, Rol, Usuario
        rol = Rol.objects.create(nombre='Mostrador service test')
        rol.permisos.set(Permiso.objects.filter(codigo='ver_inventario'))
        self.empleado = Usuario.objects.create_user(
            email='svc@celtuc.test', username='svc.inv', password='x', rol=rol,
        )
        self.api = APIClient()
        self.api.force_authenticate(self.empleado)

    def test_service_cobra_sin_tocar_stock(self):
        venta = registrar_venta(self.solar, [{
            'tipo': 'service',
            'descripcion': 'Baterias · iPhone 13 · Original',
            'item_service': self.item_service,
            'cantidad': 1,
            'precio_unitario': Decimal('45000'),
        }], usuario=self.empleado)
        self.assertEqual(venta.total, Decimal('45000'))
        self.assertEqual(MovimientoStock.objects.filter(tipo=MovimientoStock.Tipo.VENTA).count(), 0)
        item = venta.items.get()
        self.assertIsNone(item.producto_id)
        self.assertEqual(item.item_service_id, self.item_service.pk)
        self.assertEqual(item.detalle, 'Baterias · iPhone 13 · Original')

    def test_venta_mixta_descuenta_solo_la_mercaderia(self):
        r = self.api.post('/api/inventario/ventas/', {
            'sucursal': self.solar.id,
            'items': [
                {'producto': self.fuente.id, 'cantidad': 2, 'precio_unitario': 10000},
                {'tipo': 'service', 'item_service': self.item_service.id,
                 'descripcion': 'Cambio de bateria', 'cantidad': 1, 'precio_unitario': 45000},
                {'tipo': 'otro', 'descripcion': 'Mano de obra', 'cantidad': 1,
                 'precio_unitario': 5000},
            ],
        }, format='json')
        self.assertEqual(r.status_code, 201)
        self.assertEqual(float(r.data['total']), 70000)
        self.assertEqual(
            StockProducto.objects.get(producto=self.fuente, sucursal=self.solar).cantidad, 8,
        )
        self.assertEqual(MovimientoStock.objects.filter(tipo=MovimientoStock.Tipo.VENTA).count(), 1)
        nombres = [i['nombre'] for i in r.data['items']]
        self.assertEqual(nombres, ['Fuente service test', 'Cambio de bateria', 'Mano de obra'])

    def test_service_sin_descripcion_es_400(self):
        r = self.api.post('/api/inventario/ventas/', {
            'sucursal': self.solar.id,
            'items': [{'tipo': 'service', 'cantidad': 1, 'precio_unitario': 45000}],
        }, format='json')
        self.assertEqual(r.status_code, 400)

    def test_producto_sin_producto_es_400(self):
        r = self.api.post('/api/inventario/ventas/', {
            'sucursal': self.solar.id,
            'items': [{'tipo': 'producto', 'descripcion': 'algo', 'cantidad': 1,
                       'precio_unitario': 100}],
        }, format='json')
        self.assertEqual(r.status_code, 400)

    def test_la_venta_de_un_service_entra_al_arqueo(self):
        """El detalle del movimiento de caja no puede romper sin producto."""
        from caja.models import Caja, abrir_caja, registrar_venta_en_caja

        # Sin cajas con canal fiscal la venta va a la unica sesion abierta (el
        # enrutamiento por canal se prueba en caja/tests.py).
        caja = Caja.objects.create(nombre='Caja service test')
        Caja.todos.exclude(pk=caja.pk).delete()
        abrir_caja(caja, fondo_inicial=0)
        venta = registrar_venta(self.solar, [{
            'tipo': 'service', 'descripcion': 'Cambio de modulo', 'cantidad': 1,
            'precio_unitario': Decimal('120000'),
        }])
        movimientos, _ = registrar_venta_en_caja(venta, caja=caja)
        self.assertEqual(len(movimientos), 1)
        self.assertEqual(movimientos[0].monto, Decimal('120000'))
        self.assertIn('Cambio de modulo', movimientos[0].detalle)


class ApiInventarioTests(TestCase):
    """Permisos y contratos de la API."""

    def setUp(self):
        self.producto = _producto()
        self.solar = Sucursal.objects.create(nombre='Solar test', orden=1)
        self.centro = Sucursal.objects.create(nombre='Centro test', orden=2)

        self.admin = Usuario.objects.create_superuser(
            email='admin@celtuc.test', username='admin.inv', password='x',
        )
        # Empleado CON ver_inventario (rol propio para no depender del seed).
        rol = Rol.objects.create(nombre='Mostrador test')
        rol.permisos.set(Permiso.objects.filter(codigo='ver_inventario'))
        self.empleado = Usuario.objects.create_user(
            email='empleado@celtuc.test', username='empleado.inv', password='x', rol=rol,
        )
        # Usuario sin ningun permiso.
        self.pelado = Usuario.objects.create_user(
            email='pelado@celtuc.test', username='pelado.inv', password='x',
        )

    def _cliente(self, usuario):
        cliente = APIClient()
        cliente.force_authenticate(usuario)
        return cliente

    def test_empleado_con_permiso_lee_y_ajusta(self):
        cliente = self._cliente(self.empleado)
        r = cliente.get('/api/inventario/stock/')
        self.assertEqual(r.status_code, 200)
        r = cliente.post('/api/inventario/stock/ajustar/', {
            'producto': self.producto.id, 'sucursal': self.solar.id, 'delta': 3,
        }, format='json')
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.data['stock']['cantidad'], 3)
        self.assertIs(r.data['stock']['sin_dato'], False)
        self.assertEqual(r.data['movimiento']['tipo'], 'ingreso')
        self.assertEqual(r.data['movimiento']['usuario'], 'empleado.inv')

    def test_sin_permiso_403(self):
        cliente = self._cliente(self.pelado)
        self.assertEqual(cliente.get('/api/inventario/stock/').status_code, 403)
        r = cliente.post('/api/inventario/stock/ajustar/', {
            'producto': self.producto.id, 'sucursal': self.solar.id, 'delta': 1,
        }, format='json')
        self.assertEqual(r.status_code, 403)

    def test_sucursales_solo_admin_escribe(self):
        empleado = self._cliente(self.empleado)
        self.assertEqual(empleado.get('/api/inventario/sucursales/').status_code, 200)
        r = empleado.post('/api/inventario/sucursales/', {'nombre': 'Norte'}, format='json')
        self.assertEqual(r.status_code, 403)
        admin = self._cliente(self.admin)
        r = admin.post('/api/inventario/sucursales/', {'nombre': 'Norte'}, format='json')
        self.assertEqual(r.status_code, 201)

    def test_ajuste_no_baja_de_cero_da_400_legible(self):
        cliente = self._cliente(self.empleado)
        r = cliente.post('/api/inventario/stock/ajustar/', {
            'producto': self.producto.id, 'sucursal': self.solar.id, 'delta': -1,
        }, format='json')
        self.assertEqual(r.status_code, 400)
        self.assertIn('stock suficiente', r.data['detail'])

    def test_actualizar_solo_minimo_no_genera_movimiento(self):
        cliente = self._cliente(self.empleado)
        r = cliente.post('/api/inventario/stock/ajustar/', {
            'producto': self.producto.id, 'sucursal': self.solar.id, 'stock_minimo': 5,
        }, format='json')
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.data['stock']['stock_minimo'], 5)
        self.assertIsNone(r.data['movimiento'])
        self.assertEqual(MovimientoStock.objects.filter(producto=self.producto).count(), 0)

    def test_transferencia_api(self):
        cliente = self._cliente(self.empleado)
        cliente.post('/api/inventario/stock/ajustar/', {
            'producto': self.producto.id, 'sucursal': self.solar.id, 'delta': 8,
        }, format='json')
        r = cliente.post('/api/inventario/stock/transferir/', {
            'producto': self.producto.id, 'origen': self.solar.id,
            'destino': self.centro.id, 'cantidad': 3,
        }, format='json')
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.data['origen']['cantidad'], 5)
        self.assertEqual(r.data['destino']['cantidad'], 3)

    def test_movimientos_filtrables(self):
        cliente = self._cliente(self.empleado)
        cliente.post('/api/inventario/stock/ajustar/', {
            'producto': self.producto.id, 'sucursal': self.solar.id, 'delta': 2, 'nota': 'llegó caja',
        }, format='json')
        r = cliente.get(f'/api/inventario/movimientos/?producto={self.producto.id}&sucursal={self.solar.id}')
        self.assertEqual(r.status_code, 200)
        self.assertEqual(len(r.data), 1)
        self.assertEqual(r.data[0]['nota'], 'llegó caja')

    def test_requiere_autenticacion(self):
        self.assertEqual(APIClient().get('/api/inventario/stock/').status_code, 401)


class IngresoCompraventaApiTests(TestCase):
    """POST /compraventa/ingresar/: alta de un usado SIN ser admin, auditada."""

    URL = '/api/inventario/compraventa/ingresar/'

    def setUp(self):
        self.central = Sucursal.objects.create(nombre='Central test', orden=1)
        rol = Rol.objects.create(nombre='Mostrador cv test')
        rol.permisos.set(Permiso.objects.filter(codigo='ver_inventario'))
        self.empleado = Usuario.objects.create_user(
            email='cv@celtuc.test', username='empleado.cv', password='x', rol=rol,
        )

    def _cliente(self, usuario=None):
        cliente = APIClient()
        cliente.force_authenticate(usuario or self.empleado)
        return cliente

    def _payload(self, **extra):
        base = {
            'marca': 'Apple', 'modelo': 'iPhone 17 Pro', 'color': 'Negro',
            'imei1': '356938035643809', 'imei2': '356938035643810',
            'cupon': 'CV-001', 'bateria': 99, 'sucursal': self.central.id,
        }
        base.update(extra)
        return base

    def test_empleado_sin_admin_da_de_alta_con_auditoria_completa(self):
        r = self._cliente().post(self.URL, self._payload(), format='json')
        self.assertEqual(r.status_code, 201)
        self.assertFalse(r.data['reutilizado'])

        producto = Producto.objects.get(pk=r.data['producto']['id'])
        self.assertEqual(producto.nombre, 'Apple iPhone 17 Pro (usado) · 99% bat.')
        self.assertEqual(producto.calidad, 'Usado')
        self.assertIn('IMEI 356938035643809 / 356938035643810', producto.nota)
        self.assertIn('Negro', producto.nota)
        self.assertIn('Cupón CV-001', producto.nota)
        self.assertEqual(producto.categoria.nombre, 'Equipos usados')
        self.assertTrue(producto.categoria.es_equipo)

        # Auditoria de ModeloBase: quien creo cada cosa.
        self.assertEqual(producto.creado_por, self.empleado)
        self.assertEqual(producto.categoria.creado_por, self.empleado)

        self.assertEqual(r.data['stock']['cantidad'], 1)
        movimiento = MovimientoStock.objects.get(producto=producto)
        self.assertEqual(movimiento.tipo, MovimientoStock.Tipo.INGRESO)
        self.assertEqual(movimiento.creado_por, self.empleado)
        self.assertIn('cupón CV-001', movimiento.nota)

        # Auditoria por señales: un registro 'crear' por cada modelo tocado.
        from auditoria.models import RegistroAuditoria
        modelos = set(
            RegistroAuditoria.objects
            .filter(usuario=self.empleado, accion='crear')
            .values_list('modelo', flat=True)
        )
        self.assertIn(str(Producto._meta.verbose_name), modelos)
        self.assertIn(str(CategoriaProducto._meta.verbose_name), modelos)
        self.assertIn(str(MovimientoStock._meta.verbose_name), modelos)

    def test_mismo_imei_no_duplica_el_producto(self):
        cliente = self._cliente()
        cliente.post(self.URL, self._payload(), format='json')
        r = cliente.post(self.URL, self._payload(), format='json')
        self.assertEqual(r.status_code, 201)
        self.assertTrue(r.data['reutilizado'])
        self.assertEqual(Producto.objects.filter(nota__contains='356938035643809').count(), 1)
        self.assertEqual(r.data['stock']['cantidad'], 2)

    def test_categoria_usados_se_reutiliza(self):
        cliente = self._cliente()
        cliente.post(self.URL, self._payload(), format='json')
        cliente.post(self.URL, self._payload(imei1='999990001112223'), format='json')
        self.assertEqual(
            CategoriaProducto.objects.filter(nombre__icontains='usad').count(), 1,
        )

    def test_requiere_marca_o_modelo(self):
        r = self._cliente().post(self.URL, self._payload(marca='  ', modelo=''), format='json')
        self.assertEqual(r.status_code, 400)

    def test_sin_permiso_403(self):
        pelado = Usuario.objects.create_user(
            email='pelado.cv@celtuc.test', username='pelado.cv', password='x',
        )
        r = self._cliente(pelado).post(self.URL, self._payload(), format='json')
        self.assertEqual(r.status_code, 403)
        # El catalogo viene sembrado por las migraciones: alcanza con verificar
        # que el equipo del contrato no se dio de alta.
        self.assertFalse(Producto.objects.filter(nota__contains='356938035643809').exists())


class CostoSoloAdminTests(TestCase):
    """`costo_usd` de Producto: visible para admins, oculto para el resto."""

    def setUp(self):
        self.producto = _producto(costo_usd=Decimal('10.5'))
        self.admin = Usuario.objects.create_superuser(
            email='admin2@celtuc.test', username='admin.costo', password='x',
        )
        rol = Rol.objects.create(nombre='Vendedor test')
        rol.permisos.set(Permiso.objects.filter(codigo__in=['ver_productos', 'ver_inventario']))
        self.empleado = Usuario.objects.create_user(
            email='vende@celtuc.test', username='vende.costo', password='x', rol=rol,
        )

    def _get(self, usuario):
        cliente = APIClient()
        cliente.force_authenticate(usuario)
        r = cliente.get('/api/productos/items/')
        self.assertEqual(r.status_code, 200)
        return next(p for p in r.data if p['id'] == self.producto.id)

    def test_admin_ve_costo(self):
        self.assertEqual(self._get(self.admin)['costo_usd'], 10.5)

    def test_empleado_no_ve_costo(self):
        self.assertNotIn('costo_usd', self._get(self.empleado))

    def test_ver_inventario_habilita_leer_productos(self):
        rol = Rol.objects.create(nombre='Solo inventario test')
        rol.permisos.set(Permiso.objects.filter(codigo='ver_inventario'))
        usuario = Usuario.objects.create_user(
            email='soloinv@celtuc.test', username='solo.inv', password='x', rol=rol,
        )
        cliente = APIClient()
        cliente.force_authenticate(usuario)
        self.assertEqual(cliente.get('/api/productos/items/').status_code, 200)
        self.assertEqual(cliente.get('/api/productos/categorias/').status_code, 200)


class SeedInventarioTests(TestCase):
    """Los seeds de sucursales y stock inicial importado de las planillas."""

    def test_sucursales_seed(self):
        # Tras la unificación (0012) los locales quedan con su nombre definitivo:
        # Solar -> Solar YB, Centro -> Salta, más Central YB (sin stock aún).
        nombres = set(Sucursal.objects.values_list('nombre', flat=True))
        self.assertIn('Solar YB', nombres)
        self.assertIn('Salta', nombres)
        self.assertIn('Central YB', nombres)

    def test_stock_importado(self):
        solar = Sucursal.objects.get(nombre='Solar YB')
        centro = Sucursal.objects.get(nombre='Salta')
        self.assertGreater(StockProducto.objects.filter(sucursal=solar, cantidad__gt=0).count(), 150)
        self.assertGreater(StockProducto.objects.filter(sucursal=centro, cantidad__gt=0).count(), 130)
        # Cada fila importada dejo su movimiento de carga inicial.
        self.assertEqual(
            MovimientoStock.objects.filter(nota__icontains='planilla').count(),
            StockProducto.objects.filter(cantidad__gt=0).count(),
        )

    def test_seed_no_informado(self):
        # Las filas cuya celda de stock estaba vacia en las planillas quedan
        # como "(no informado)": cantidad 0 + sin_dato, sin movimiento.
        # 378 del seed 0006 (Solar 175 + Centro 203) + 506 repuestos de
        # service del seed 0007 (253 productos x 2 sucursales) + 86 iPhones
        # del seed 0008 (43 modelos x 2 sucursales).
        marcadas = StockProducto.objects.filter(sin_dato=True)
        self.assertEqual(marcadas.count(), 970)
        self.assertFalse(marcadas.exclude(cantidad=0).exists())
        # Producto que solo estaba en las planillas sin cantidad: se creo aca.
        haylou = Producto.objects.get(nombre='Haylou X1 Neo')
        self.assertEqual(haylou.stocks.filter(sin_dato=True).count(), 2)
        # "Fuente 5W - CO" estaba vacia en las dos hojas (en Centro ademas de
        # Solar): las dos sucursales quedan "(no informado)".
        fuente5 = StockProducto.objects.filter(
            producto__nombre='Fuente 5W', producto__calidad='Calidad original',
        )
        self.assertEqual(fuente5.count(), 2)
        self.assertTrue(all(f.sin_dato and f.cantidad == 0 for f in fuente5))
        # Los repuestos de service (seed 0007) tambien: la bateria del 14 Pro
        # existe como producto SIN precio (los precios siguen en /service).
        bateria = Producto.objects.get(categoria__nombre='Baterías', nombre='14 PRO')
        self.assertIsNone(bateria.precio_lista_usd)
        self.assertEqual(bateria.stocks.filter(sin_dato=True, cantidad=0).count(), 2)
        # Los iPhone 17 de la planilla no tienen precio $ cargado -> no entran.
        self.assertFalse(Producto.objects.filter(
            categoria__nombre='Baterías', nombre__startswith='17',
        ).exists())
        # Los equipos iPhone (seed 0008): 43 modelos sin precio, vinculados a
        # su Dispositivo (para la Ficha) y "(no informado)" en ambas sucursales.
        iphones = Producto.objects.filter(categoria__nombre='iPhones')
        self.assertEqual(iphones.count(), 43)
        quince_pro = iphones.get(nombre='iPhone 15 Pro')
        self.assertIsNone(quince_pro.precio_lista_usd)
        self.assertEqual(quince_pro.stocks.filter(sin_dato=True, cantidad=0).count(), 2)
        self.assertEqual(
            [d.nombre for d in quince_pro.dispositivos.all()], ['iPhone 15 Pro'],
        )

    def test_casos_conocidos_de_la_hoja(self):
        # Las variables conservan el nombre de las planillas de origen (hoja
        # "Solar" y hoja "Centro"), hoy renombradas a Solar YB y Salta.
        solar = Sucursal.objects.get(nombre='Solar YB')
        centro = Sucursal.objects.get(nombre='Salta')
        # "Fuente 20W - CO" -> 56 en Solar y 32 en Centro (celdas I5).
        fila = StockProducto.objects.get(
            producto__nombre='Fuente 20W', producto__calidad='Calidad original',
            producto__nota='', sucursal=solar,
        )
        self.assertEqual(fila.cantidad, 56)
        fila_centro = StockProducto.objects.get(producto=fila.producto, sucursal=centro)
        self.assertEqual(fila_centro.cantidad, 32)
        # Los tipeados a mano al final de la hoja Solar existen como productos.
        self.assertTrue(Producto.objects.filter(nombre__iexact='alexa echo show 10').exists())


# ===== Importacion de stock por sucursal =====

def _planilla(filas, encabezado=None):
    """Un .xlsx en memoria con la forma de la planilla del negocio.

    `filas` son tuplas (seccion, producto, lista_usd, stock, minimo); un None en
    `producto` genera una fila de sub-encabezado (la que en MODULOS reusa la
    columna STOCK para los precios CO/AO).
    """
    import io

    import openpyxl

    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.append(encabezado or [
        None, 'PRODUCTOS', 'COSTO USD', 'COSTO $', 'PRECIO DE LISTA USD',
        'PRECIO CASH USD (20% OFF)', 'PRECIO DE LISTA CREDITO', '$/DEBITO/TRANSF',
        'STOCK', 'STOCK MINIMO',
    ])
    for seccion, producto, lista, stock, minimo in filas:
        hoja.append([seccion, producto, None, None, lista, None, None, None, stock, minimo])
    buffer = io.BytesIO()
    libro.save(buffer)
    buffer.seek(0)
    buffer.name = 'StockPorSucursal.xlsx'
    return buffer


class PlanillaImportacionTests(TestCase):
    """La lectura del archivo: que fila es un producto y que celda es un conteo."""

    def setUp(self):
        from .importacion import analizar

        self.analizar = analizar
        self.sucursal = Sucursal.objects.create(nombre='Import test', orden=90)
        categoria = CategoriaProducto.objects.create(nombre='Zetatest accesorios')
        self.producto = Producto.objects.create(
            categoria=categoria, nombre='Zetatest Cable 1M',
            calidad='Calidad original', precio_lista_usd=Decimal('10'),
        )

    def _filas(self, planilla):
        return {f['fila']: f for f in self.analizar(planilla, self.sucursal)['filas']}

    def test_celda_vacia_no_es_cero(self):
        """Sin cantidad en la planilla el stock NO se toca (si no, borraria todo)."""
        filas = self._filas(_planilla([
            ('ZETATEST ACCESORIOS', 'Zetatest Cable 1M - CO', 10, None, None),
        ]))
        fila = filas[2]
        self.assertEqual(fila['estado'], 'sin_valor')
        self.assertFalse(fila['sugerido'])
        self.assertIsNone(fila['cantidad_nueva'])

    def test_cero_explicito_si_se_importa(self):
        filas = self._filas(_planilla([
            ('ZETATEST ACCESORIOS', 'Zetatest Cable 1M - CO', 10, 0, None),
        ]))
        self.assertEqual(filas[2]['estado'], 'actualiza')
        self.assertEqual(filas[2]['cantidad_nueva'], 0)

    def test_valor_que_no_es_conteo_queda_invalido(self):
        filas = self._filas(_planilla([
            ('ZETATEST ACCESORIOS', 'Zetatest Cable 1M - CO', 10, 88800, None),
        ]))
        self.assertEqual(filas[2]['estado'], 'invalida')
        self.assertIn('precio', filas[2]['motivo'])

    def test_texto_de_error_de_excel_queda_invalido(self):
        filas = self._filas(_planilla([
            ('ZETATEST ACCESORIOS', 'Zetatest Cable 1M - CO', 10, '#VALUE!', None),
        ]))
        self.assertEqual(filas[2]['estado'], 'invalida')

    def test_seccion_que_reusa_la_columna_stock(self):
        """Como MODULOS: un sub-encabezado avisa que ahi van precios, no unidades."""
        filas = self._filas(_planilla([
            ('ZETATEST ACCESORIOS', 'Zetatest Cable 1M - CO', 10, 3, None),
            (None, None, None, 'CO', 'AO'),
            ('MODULOS ZETATEST', 'Zetatest Cable 1M - CO', 10, 5, None),
        ]))
        self.assertEqual(filas[2]['estado'], 'actualiza')
        self.assertEqual(filas[4]['estado'], 'invalida')
        self.assertIn('CO/AO', filas[4]['motivo'])

    def test_sin_encabezado_error_legible(self):
        from django.core.exceptions import ValidationError

        planilla = _planilla([('X', 'Y', 1, 1, None)], encabezado=['a', 'b', 'c'])
        with self.assertRaises(ValidationError) as ctx:
            self.analizar(planilla, self.sucursal)
        self.assertIn('PRODUCTOS', ' '.join(ctx.exception.messages))


class MatcheoImportacionTests(TestCase):
    """El cruce con el catalogo: exacto, aproximado y lo que se manda a revisar."""

    def setUp(self):
        from .importacion import analizar

        self.analizar = analizar
        self.sucursal = Sucursal.objects.create(nombre='Match test', orden=91)
        self.categoria = CategoriaProducto.objects.create(nombre='Zetatest cables')
        self.cable = Producto.objects.create(
            categoria=self.categoria, nombre='Zetatest Cable 1M',
            calidad='Calidad original', precio_lista_usd=Decimal('10'),
        )

    def _fila(self, nombre, stock=1, seccion='ZETATEST CABLES', lista=10):
        planilla = _planilla([(seccion, nombre, lista, stock, None)])
        return self.analizar(planilla, self.sucursal)['filas'][0]

    def test_nombre_con_calidad_abreviada(self):
        """La planilla escribe "- CO" lo que el catalogo guarda como calidad."""
        fila = self._fila('Zetatest Cable 1M - CO')
        self.assertEqual(fila['producto'], self.cable.id)
        self.assertEqual(fila['confianza'], 'exacta')

    def test_diferencia_de_numero_no_es_el_mismo_producto(self):
        """256GB y 512GB no son el mismo producto por mas parecido que sea el texto."""
        Producto.objects.create(
            categoria=self.categoria, nombre='Zetatest Tablet 512GB',
            precio_lista_usd=Decimal('100'),
        )
        fila = self._fila('Zetatest Tablet 256GB', lista=100)
        self.assertIsNone(fila['producto'])
        self.assertEqual(fila['estado'], 'nueva')

    def test_erratas_dan_coincidencia_aproximada(self):
        fila = self._fila('Zetatest Cable 1M - CO ( suelto )')
        self.assertEqual(fila['producto'], self.cable.id)
        self.assertEqual(fila['confianza'], 'aproximada')

    def test_dos_productos_con_el_mismo_nombre_van_a_revisar(self):
        Producto.objects.create(
            categoria=self.categoria, nombre='Zetatest Cable 1M',
            calidad='Calidad original', precio_lista_usd=Decimal('99'),
        )
        fila = self._fila('Zetatest Cable 1M - CO', lista=None)
        self.assertEqual(fila['estado'], 'revisar')
        self.assertEqual(len(fila['candidatos']), 2)
        self.assertFalse(fila['sugerido'])

    def test_el_precio_de_lista_desempata(self):
        otro = Producto.objects.create(
            categoria=self.categoria, nombre='Zetatest Cable 1M',
            calidad='Calidad original', precio_lista_usd=Decimal('99'),
        )
        fila = self._fila('Zetatest Cable 1M - CO', lista=99)
        self.assertEqual(fila['estado'], 'actualiza')
        self.assertEqual(fila['producto'], otro.id)

    def test_producto_que_no_esta_en_el_catalogo(self):
        fila = self._fila('Zetatest Soporte Magnetico XYZ')
        self.assertEqual(fila['estado'], 'nueva')
        self.assertTrue(fila['puede_crear'])
        self.assertEqual(fila['categoria_id'], self.categoria.id)

    def test_antes_y_despues_de_una_fila_existente(self):
        aplicar_ajuste(self.cable, self.sucursal, delta=4)
        fila = self._fila('Zetatest Cable 1M - CO', stock=9)
        self.assertEqual(fila['cantidad_actual'], 4)
        self.assertEqual(fila['cantidad_nueva'], 9)
        self.assertEqual(fila['estado'], 'actualiza')
        self.assertTrue(fila['sugerido'])

    def test_misma_cantidad_no_propone_nada(self):
        aplicar_ajuste(self.cable, self.sucursal, cantidad=4)
        fila = self._fila('Zetatest Cable 1M - CO', stock=4)
        self.assertEqual(fila['estado'], 'igual')
        self.assertFalse(fila['sugerido'])

    def test_dos_filas_al_mismo_producto_quedan_marcadas(self):
        """La planilla puede ser mas fina que el catalogo ("8" y "8+" = "8 / 8+")."""
        combinado = Producto.objects.create(
            categoria=self.categoria, nombre='Zetatest Modulo 8 / 8+',
            precio_lista_usd=Decimal('50'),
        )
        planilla = _planilla([
            ('ZETATEST CABLES', 'Zetatest Modulo 8 / 8+', 50, 3, None),
            ('ZETATEST CABLES', 'Zetatest Modulo 8 / 8+', 50, 9, None),
        ])
        filas = self.analizar(planilla, self.sucursal)['filas']
        self.assertEqual([f['producto'] for f in filas], [combinado.id, combinado.id])
        self.assertEqual(filas[0]['duplicada_con'], [3])
        self.assertEqual(filas[1]['duplicada_con'], [2])
        # Ninguna viene marcada: aplicar las dos pisaria una con la otra.
        self.assertFalse(any(f['sugerido'] for f in filas))
        self.assertIn('mismo producto', filas[0]['motivo'])


class ImportarStockApiTests(TestCase):
    """Los dos endpoints: analizar (no escribe) y aplicar (todo o nada)."""

    ANALIZAR = '/api/inventario/stock/importar/analizar/'
    APLICAR = '/api/inventario/stock/importar/aplicar/'

    def setUp(self):
        self.sucursal = Sucursal.objects.create(nombre='Api import test', orden=92)
        self.categoria = CategoriaProducto.objects.create(nombre='Zetatest api')
        self.producto = Producto.objects.create(
            categoria=self.categoria, nombre='Zetatest Api Cable',
            precio_lista_usd=Decimal('12'),
        )
        self.admin = Usuario.objects.create_superuser(
            email='admin@import.test', username='admin.import', password='x',
        )
        rol = Rol.objects.create(nombre='Mostrador import test')
        rol.permisos.set(Permiso.objects.filter(codigo='ver_inventario'))
        self.empleado = Usuario.objects.create_user(
            email='emp@import.test', username='empleado.import', password='x', rol=rol,
        )
        self.pelado = Usuario.objects.create_user(
            email='nada@import.test', username='pelado.import', password='x',
        )

    def _cliente(self, usuario):
        cliente = APIClient()
        cliente.force_authenticate(usuario)
        return cliente

    def _subir(self, usuario, filas):
        return self._cliente(usuario).post(self.ANALIZAR, {
            'sucursal': self.sucursal.id, 'archivo': _planilla(filas),
        }, format='multipart')

    def test_analizar_devuelve_el_diff_y_no_escribe(self):
        r = self._subir(self.empleado, [('ZETATEST API', 'Zetatest Api Cable', 12, 7, 2)])
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.data['sucursal'], self.sucursal.id)
        fila = r.data['filas'][0]
        self.assertEqual(fila['producto'], self.producto.id)
        self.assertEqual(fila['cantidad_nueva'], 7)
        self.assertEqual(fila['minimo_nuevo'], 2)
        self.assertEqual(r.data['resumen']['actualiza'], 1)
        # Analizar es de solo lectura.
        self.assertEqual(StockProducto.objects.filter(sucursal=self.sucursal).count(), 0)

    def test_archivo_que_no_es_xlsx(self):
        import io

        archivo = io.BytesIO(b'no soy un excel')
        archivo.name = 'stock.csv'
        r = self._cliente(self.empleado).post(self.ANALIZAR, {
            'sucursal': self.sucursal.id, 'archivo': archivo,
        }, format='multipart')
        self.assertEqual(r.status_code, 400)
        self.assertIn('.xlsx', str(r.data))

    def test_sin_permiso_403(self):
        r = self._subir(self.pelado, [('ZETATEST API', 'Zetatest Api Cable', 12, 7, None)])
        self.assertEqual(r.status_code, 403)

    def test_requiere_autenticacion(self):
        self.assertEqual(APIClient().post(self.ANALIZAR, {}).status_code, 401)

    def test_aplicar_fija_stock_y_deja_kardex(self):
        cliente = self._cliente(self.empleado)
        r = cliente.post(self.APLICAR, {
            'sucursal': self.sucursal.id,
            'archivo': 'StockPorSucursal.xlsx',
            'items': [{'producto': self.producto.id, 'cantidad': 7, 'stock_minimo': 2}],
        }, format='json')
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.data['actualizados'], 1)
        self.assertEqual(r.data['creados'], 0)
        fila = StockProducto.objects.get(producto=self.producto, sucursal=self.sucursal)
        self.assertEqual(fila.cantidad, 7)
        self.assertEqual(fila.stock_minimo, 2)
        self.assertFalse(fila.sin_dato)
        movimiento = MovimientoStock.objects.get(producto=self.producto, sucursal=self.sucursal)
        self.assertEqual(movimiento.delta, 7)
        self.assertIn('StockPorSucursal.xlsx', movimiento.nota)
        self.assertEqual(movimiento.creado_por, self.empleado)

    def test_aplicar_sin_items_da_400(self):
        r = self._cliente(self.empleado).post(self.APLICAR, {
            'sucursal': self.sucursal.id, 'items': [],
        }, format='json')
        self.assertEqual(r.status_code, 400)

    def test_crear_productos_es_solo_admin(self):
        alta = {
            'crear': {'nombre': 'Zetatest Nuevo', 'categoria': self.categoria.id, 'lista_usd': '15'},
            'cantidad': 3,
        }
        r = self._cliente(self.empleado).post(self.APLICAR, {
            'sucursal': self.sucursal.id, 'items': [alta],
        }, format='json')
        self.assertEqual(r.status_code, 403)
        self.assertFalse(Producto.objects.filter(nombre='Zetatest Nuevo').exists())

        r = self._cliente(self.admin).post(self.APLICAR, {
            'sucursal': self.sucursal.id, 'items': [alta],
        }, format='json')
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.data['creados'], 1)
        creado = Producto.objects.get(nombre='Zetatest Nuevo')
        self.assertEqual(creado.categoria, self.categoria)
        self.assertEqual(creado.precio_lista_usd, Decimal('15'))
        self.assertEqual(creado.stocks.get(sucursal=self.sucursal).cantidad, 3)

    def test_una_fila_invalida_no_aplica_ninguna(self):
        """Todo o nada: el stock no queda a medio importar."""
        r = self._cliente(self.empleado).post(self.APLICAR, {
            'sucursal': self.sucursal.id,
            'items': [
                {'producto': self.producto.id, 'cantidad': 5},
                {'producto': 999999, 'cantidad': 2},
            ],
        }, format='json')
        self.assertEqual(r.status_code, 400)
        self.assertEqual(StockProducto.objects.filter(sucursal=self.sucursal).count(), 0)

    def test_producto_repetido_da_400(self):
        r = self._cliente(self.empleado).post(self.APLICAR, {
            'sucursal': self.sucursal.id,
            'items': [
                {'producto': self.producto.id, 'cantidad': 5},
                {'producto': self.producto.id, 'cantidad': 9},
            ],
        }, format='json')
        self.assertEqual(r.status_code, 400)


class ExportarEnFormatoImportadorTests(TestCase):
    """El archivo que baja «Exportar inventario» con «Mismo formato que
    importador» tiene que volver a entrar por «Importar stock» sin tocar nada.

    El generador vive en el front (`exportar/xlsx.ts`), asi que lo que se prueba
    aca es el CONTRATO que ese archivo tiene que cumplir: los rotulos de la fila
    1, la categoria en la columna A (combinada, o sea con valor solo en la
    primera fila del grupo), el nombre rearmado como "nombre + calidad + nota" y
    la regla de que una fila sin contar viaja VACIA. Si alguien cambia el parser
    o los rotulos, este test se cae antes de que se rompa el ida y vuelta.
    """

    # Los rotulos EXACTOS que escribe el exportador en modo importador
    # (`COLUMNAS_IMPORTADOR` en `frontend/src/components/inventario/exportar/datos.ts`).
    ENCABEZADO = [
        None, 'PRODUCTOS', 'COSTO USD', 'COSTO $', 'PRECIO DE LISTA USD',
        'PRECIO CASH USD (20% OFF)',
        'PRECIO DE LISTA CREDITO 1-3 CUOTAS  SIN INTERES',
        '$/DEBITO/TRANSF (20% OFF)', 'STOCK', 'STOCK MINIMO',
    ]

    def setUp(self):
        from .importacion import analizar

        self.analizar = analizar
        self.sucursal = Sucursal.objects.create(nombre='Export ida y vuelta', orden=91)
        self.categoria = CategoriaProducto.objects.create(nombre='Zetaexport fuentes')
        # Un producto con calidad y otro sin nada: los dos casos del rearmado.
        self.con_calidad = Producto.objects.create(
            categoria=self.categoria, nombre='Zetaexport Fuente 20W',
            calidad='Calidad original', precio_lista_usd=Decimal('25'),
        )
        self.simple = Producto.objects.create(
            categoria=self.categoria, nombre='Zetaexport Cable 2M',
            precio_lista_usd=Decimal('8'),
        )
        aplicar_ajuste(self.con_calidad, self.sucursal, delta=7)

    def _archivo_del_exportador(self, filas):
        """Reproduce el layout que genera el exportador en modo importador."""
        import io

        import openpyxl

        libro = openpyxl.Workbook()
        hoja = libro.active
        hoja.append(self.ENCABEZADO)
        for categoria, nombre, lista, stock, minimo in filas:
            hoja.append([categoria, nombre, None, None, lista, None, None, None, stock, minimo])
        # La nota que el exportador deja en la columna K: NO tiene que leerse
        # como un renglon de producto.
        hoja.cell(1, 11).value = 'Stock de Export ida y vuelta. Completa la columna STOCK y volve a subirlo.'
        buffer = io.BytesIO()
        libro.save(buffer)
        buffer.seek(0)
        buffer.name = 'inventario-export.xlsx'
        return buffer

    def test_el_archivo_exportado_vuelve_a_entrar_sin_revisar(self):
        # Asi lo escribe el exportador: la categoria solo en la primera fila del
        # grupo (el resto son celdas de la combinacion) y el nombre completo.
        archivo = self._archivo_del_exportador([
            ('Zetaexport fuentes', 'Zetaexport Fuente 20W Calidad original', 25, 7, None),
            (None, 'Zetaexport Cable 2M', 8, None, None),
        ])
        resultado = self.analizar(archivo, self.sucursal)
        filas = {f['fila']: f for f in resultado['filas']}

        # Dos renglones: la nota de la columna K no entra como producto.
        self.assertEqual(len(resultado['filas']), 2)

        # El que tenia stock informado matchea y se ve IGUAL a lo que ya hay.
        self.assertEqual(filas[2]['estado'], 'igual')
        self.assertEqual(filas[2]['seccion'], 'Zetaexport fuentes')

        # El que viajo sin contar no toca el stock (vacio no es cero).
        self.assertEqual(filas[3]['estado'], 'sin_valor')
        self.assertIsNone(filas[3]['cantidad_nueva'])

        # Nada quedo para revisar ni se propuso crear un producto nuevo.
        self.assertEqual(resultado['resumen']['revisar'], 0)
        self.assertEqual(resultado['resumen']['nueva'], 0)
        self.assertEqual(resultado['resumen']['invalida'], 0)

    def test_el_conteo_corregido_a_mano_se_aplica(self):
        """El uso real: se baja el archivo, se corrige la columna STOCK y se sube."""
        archivo = self._archivo_del_exportador([
            ('Zetaexport fuentes', 'Zetaexport Fuente 20W Calidad original', 25, 12, None),
        ])
        resultado = self.analizar(archivo, self.sucursal)
        fila = resultado['filas'][0]
        self.assertEqual(fila['estado'], 'actualiza')
        self.assertEqual(fila['cantidad_actual'], 7)
        self.assertEqual(fila['cantidad_nueva'], 12)
        self.assertTrue(fila['sugerido'])  # entra marcada: cambia y no hay dudas
        self.assertEqual(resultado['resumen']['sube'], 1)

    def test_los_rotulos_del_exportador_son_los_que_busca_el_parser(self):
        """Si alguien renombra una columna del exportador, esto se cae."""
        from .importacion import _indice_encabezado

        indice, columnas = _indice_encabezado([self.ENCABEZADO])
        self.assertEqual(indice, 0)
        self.assertEqual(columnas['producto'], 1)   # columna B
        self.assertEqual(columnas['stock'], 8)      # columna I
        self.assertEqual(columnas['minimo'], 9)     # columna J
        self.assertEqual(columnas['lista'], 4)      # columna E
